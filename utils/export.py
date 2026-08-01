import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def export_csv(logs):
    """
    Generates a CSV string from a list of DailyLog objects.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    from models.question import QuestionConfig
    questions = QuestionConfig.query.order_by(QuestionConfig.id).all()
    q_map = {q.id: q.short_title for q in questions}
    
    # Write header
    writer.writerow([
        'Date', 
        f"{q_map.get(1, 'Study')} (Value)", f"{q_map.get(1, 'Study')} (Note)", 
        f"{q_map.get(2, 'Project')} (Value)", f"{q_map.get(2, 'Project')} (Note)", 
        f"{q_map.get(3, 'Exercise')} (Value)", f"{q_map.get(3, 'Exercise')} (Note)", 
        f"{q_map.get(4, 'Career')} (Value)", f"{q_map.get(4, 'Career')} (Note)", 
        f"{q_map.get(5, 'Avoided Social Media')} (Value)", f"{q_map.get(5, 'Avoided Social Media')} (Note)", 
        'Score', 'Completed'
    ])
    
    # Write rows
    for log in logs:
        writer.writerow([
            log.date.strftime('%Y-%m-%d'),
            'Yes' if log.q1_val else 'No', log.q1_note or '',
            'Yes' if log.q2_val else 'No', log.q2_note or '',
            'Yes' if log.q3_val else 'No', log.q3_note or '',
            'Yes' if log.q4_val else 'No', log.q4_note or '',
            'Yes' if log.q5_val else 'No', log.q5_note or '',
            log.score,
            'Yes' if log.completed else 'No'
        ])
        
    return output.getvalue()

def export_excel(logs):
    """
    Generates a beautifully styled Excel binary stream using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "LifeTrack Log History"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Colors (Slate / Blue Theme)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    data_font = Font(name="Segoe UI", size=10)
    yes_font = Font(name="Segoe UI", size=10, bold=True, color="10B981") # Emerald Green
    no_font = Font(name="Segoe UI", size=10, color="EF4444") # Red
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Headers
    from models.question import QuestionConfig
    questions = QuestionConfig.query.order_by(QuestionConfig.id).all()
    q_map = {q.id: q.short_title for q in questions}
    
    headers = [
        'Date', 
        q_map.get(1, 'Study'), f"{q_map.get(1, 'Study')} Notes", 
        q_map.get(2, 'Project'), f"{q_map.get(2, 'Project')} Notes", 
        q_map.get(3, 'Exercise'), f"{q_map.get(3, 'Exercise')} Notes", 
        q_map.get(4, 'Career'), f"{q_map.get(4, 'Career')} Notes", 
        q_map.get(5, 'Social Media'), f"{q_map.get(5, 'Social Media')} Notes", 
        'Score', 'Completed'
    ]
    
    ws.append(headers)
    
    # Style Header Row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # Write Rows
    for r_idx, log in enumerate(logs, start=2):
        row_data = [
            log.date.strftime('%Y-%m-%d'),
            'Yes' if log.q1_val else 'No', log.q1_note or '',
            'Yes' if log.q2_val else 'No', log.q2_note or '',
            'Yes' if log.q3_val else 'No', log.q3_note or '',
            'Yes' if log.q4_val else 'No', log.q4_note or '',
            'Yes' if log.q5_val else 'No', log.q5_note or '',
            log.score,
            'Yes' if log.completed else 'No'
        ]
        ws.append(row_data)
        
        # Style Data Row
        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Custom styling for Boolean fields
            if c_idx in [2, 4, 6, 8, 10]:
                cell.alignment = align_center
                if cell.value == 'Yes':
                    cell.font = yes_font
                else:
                    cell.font = no_font
            elif c_idx in [1, 12, 13]:
                cell.alignment = align_center
                if c_idx == 13 and cell.value == 'Yes':
                    cell.font = yes_font
            else:
                cell.alignment = align_left
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_pdf(logs, stats, report_data):
    """
    Generates a professional executive PDF report using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter, 
        rightMargin=40, leftMargin=40, 
        topMargin=40, bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom PDF Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=26,
        leading=30,
        textColor=colors.HexColor('#0F172A'), # Slate 900
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748B'), # Slate 500
        spaceAfter=30
    )
    
    h2_style = ParagraphStyle(
        'H2Style',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=15,
        textColor=colors.HexColor('#334155') # Slate 700
    )
    
    th_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    td_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155'),
        alignment=1 # Center
    )
    
    td_left_style = ParagraphStyle(
        'TableCellLeft',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155'),
        alignment=0 # Left
    )

    # Document Header
    story.append(Paragraph("LifeTrack Performance Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Offline Productivity Ledger", subtitle_style))
    story.append(Spacer(1, 10))
    
    # 1. Summary Cards Table
    story.append(Paragraph("Key Metrics Summary", h2_style))
    metric_data = [
        [
            Paragraph("<b>Current Streak</b>", body_style), 
            Paragraph("<b>Longest Streak</b>", body_style), 
            Paragraph("<b>Completion Rate</b>", body_style)
        ],
        [
            Paragraph(f"<font size=14 color='#10B981'><b>{stats['current_streak']} Days</b></font>", body_style),
            Paragraph(f"<font size=14 color='#6366F1'><b>{stats['longest_streak']} Days</b></font>", body_style),
            Paragraph(f"<font size=14 color='#F59E0B'><b>{stats['completion_percentage']}%</b></font>", body_style)
        ],
        [
            Paragraph("<b>Total Logged</b>", body_style), 
            Paragraph("<b>Avg Daily Score</b>", body_style), 
            Paragraph("<b>Consistency Level</b>", body_style)
        ],
        [
            Paragraph(f"<font size=14><b>{stats['total_logged_days']} Days</b></font>", body_style),
            Paragraph(f"<font size=14><b>{stats['avg_score']}/5.0</b></font>", body_style),
            Paragraph(f"<font size=14><b>{report_data['consistency_score']}%</b></font>", body_style)
        ]
    ]
    
    metric_table = Table(metric_data, colWidths=[170, 170, 170])
    metric_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    story.append(metric_table)
    story.append(Spacer(1, 20))
    
    # 2. Recommendations Section
    if report_data.get('recommendations'):
        story.append(Paragraph("Personalized Recommendations", h2_style))
        rec_story = []
        for rec in report_data['recommendations']:
            rec_story.append(Paragraph(f"• {rec}", body_style))
            rec_story.append(Spacer(1, 4))
        story.append(KeepTogether(rec_story))
        story.append(Spacer(1, 20))
        
    # 3. Log History Table
    story.append(Paragraph("Recent Check-in Logs", h2_style))
    
    from models.question import QuestionConfig
    questions = QuestionConfig.query.order_by(QuestionConfig.id).all()
    q_map = {q.id: q.short_title for q in questions}
    
    log_headers = [
        'Date', 
        q_map.get(1, 'Study'), 
        q_map.get(2, 'Project'), 
        q_map.get(3, 'Exercise'), 
        q_map.get(4, 'Career'), 
        q_map.get(5, 'Social Media'), 
        'Score', 
        'Completed'
    ]
    table_data = [[Paragraph(h, th_style) for h in log_headers]]
    
    # List latest 30 logs for readability in PDF (prevent excessive pages)
    pdf_logs = logs[:40] # Display up to 40 logs
    
    for log in pdf_logs:
        table_data.append([
            Paragraph(log.date.strftime('%Y-%m-%d'), td_style),
            Paragraph('Yes' if log.q1_val else 'No', td_style),
            Paragraph('Yes' if log.q2_val else 'No', td_style),
            Paragraph('Yes' if log.q3_val else 'No', td_style),
            Paragraph('Yes' if log.q4_val else 'No', td_style),
            Paragraph('Yes' if log.q5_val else 'No', td_style),
            Paragraph(f"{log.score}/5", td_style),
            Paragraph('<b>Yes</b>' if log.completed else 'No', td_style)
        ])
        
    log_table = Table(table_data, colWidths=[75, 55, 55, 55, 55, 75, 50, 70])
    log_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        # Alternate row backgrounds
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(log_table)
    
    # Build Document
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
